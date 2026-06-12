from __future__ import annotations

"""
audit_service.py
----------------
Responsible for writing the per-batch audit XLSX to disk.

Timing is owned entirely by the server:
  - Each ``AuditRecord`` carries ``ocr_start / ocr_end / llm_start / llm_end``
    as ``datetime`` objects set inside the extraction pipeline.
  - The browser never sends timing data; it only POSTs the list of filenames
    and the extraction results it received from the API.

XLSX columns per data row
-------------------------
  File Name
  Extracted Bank Name  |  Expected Bank Name
  Extracted FI Code    |  Expected FI Code
  Extracted Master Acc |  Expected Master Acc
  Extracted Sub Acc    |  Expected Sub Acc
  Accurate Count
  OCR Start  |  OCR End  |  OCR Duration
  LLM Start  |  LLM End  |  LLM Duration
"""

import logging
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

logger = logging.getLogger(__name__)

_AUDIT_DIR = Path("audit_logs").resolve()
_AUDIT_DIR.mkdir(parents=True, exist_ok=True)

_AUDIT_FIELDS = [
    ("bank_name",             "Bank Name"),
    ("fi_num",                "FI Code"),
    ("master_account_number", "Master Account No."),
    ("sub_account_number",    "Sub Account No."),
]

_TIMING_FIELDS = [
    "OCR Start", "OCR End", "OCR Duration",
    "LLM Start", "LLM End", "LLM Duration",
]


# ---------------------------------------------------------------------------
# Public data container — filled by the extraction pipeline
# ---------------------------------------------------------------------------

@dataclass
class AuditRecord:
    """
    One record per file processed in a batch run.

    All timing fields are UTC ``datetime`` objects set by the server.
    They are *not* accepted from the browser.
    """
    filename: str
    extract_result: dict | None = None   # the full API response dict, or None on error
    extract_error: str | None = None     # error message string, or None on success
    ocr_start: datetime | None = None
    ocr_end: datetime | None = None
    llm_start: datetime | None = None
    llm_end: datetime | None = None


# ---------------------------------------------------------------------------
# Private formatting helpers
# ---------------------------------------------------------------------------

def _fmt_duration(start: datetime | None, end: datetime | None) -> str:
    if start is None or end is None:
        return ""
    secs = (end - start).total_seconds()
    if secs >= 60:
        m = int(secs // 60)
        s = secs % 60
        return f"{m}m {s:.1f}s"
    return f"{secs:.1f}s"


def _fmt_ts(dt: datetime | None) -> str:
    if dt is None:
        return ""
    # Normalise to local-naive for display (strip tzinfo)
    if dt.tzinfo is not None:
        dt = dt.astimezone().replace(tzinfo=None)
    return dt.strftime("%Y-%m-%d %H:%M:%S")


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------

def write_audit_xlsx(records: list[AuditRecord]) -> Path:
    """
    Write one audit XLSX for a batch run.

    Returns the ``Path`` of the file written.
    Raises ``ImportError`` if openpyxl is not installed.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = _AUDIT_DIR / f"audit_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit"

    # ── Styles ──────────────────────────────────────────────────────────────
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    green_fill   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font   = Font(color="006100")
    red_fill     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font     = Font(color="9C0006")
    total_font   = Font(bold=True, size=11)
    batch_fill   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    batch_font   = Font(bold=True, size=11)
    thin_border  = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # ── Header row ───────────────────────────────────────────────────────────
    header = ["File Name"]
    for _key, label in _AUDIT_FIELDS:
        header += [f"Extracted {label}", f"Expected {label}"]
    header.append("Accurate Count")
    header += _TIMING_FIELDS

    ws.append(header)
    for col_idx in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Expected-column indices (1-based): columns 3, 5, 7, 9
    expected_col_indices = [2 + i * 2 + 1 for i in range(len(_AUDIT_FIELDS))]

    # ── Data rows ────────────────────────────────────────────────────────────
    total_accurate = 0
    total_fields   = 0

    for record in records:
        row_data = _build_data_row(record)
        ws.append(row_data)
        row_num = ws.max_row

        for col_idx in range(1, len(row_data) + 1):
            ws.cell(row=row_num, column=col_idx).border = thin_border

        if record.extract_error or not record.extract_result:
            # Error row — paint all expected columns red
            for col_idx in expected_col_indices:
                ws.cell(row=row_num, column=col_idx).fill = red_fill
                ws.cell(row=row_num, column=col_idx).font = red_font
            total_fields += len(_AUDIT_FIELDS)
            continue

        # Colour expected columns green/red per field match
        cmp     = (record.extract_result or {}).get("comparison") or {}
        matches = _compute_matches(record.extract_result or {})
        for i, col_idx in enumerate(expected_col_indices):
            cell = ws.cell(row=row_num, column=col_idx)
            if matches[i]:
                cell.fill = green_fill
                cell.font = green_font
            else:
                cell.fill = red_fill
                cell.font = red_font

        accurate_count = sum(matches)
        total_accurate += accurate_count
        total_fields   += len(_AUDIT_FIELDS)

    # ── Summary row ──────────────────────────────────────────────────────────
    pct = (total_accurate / total_fields * 100) if total_fields else 0.0
    summary_row = ["TOTAL"] + [""] * (len(_AUDIT_FIELDS) * 2)
    summary_row.append(f"{total_accurate}/{total_fields} ({pct:.2f}%)")
    ws.append(summary_row)
    row_num = ws.max_row
    for col_idx in range(1, len(summary_row) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font   = total_font
        cell.border = thin_border

    # ── Batch timing summary row ──────────────────────────────────────────────
    batch_ocr_start = _safe_min([r.ocr_start for r in records if r.ocr_start])
    batch_ocr_end   = _safe_max([r.ocr_end   for r in records if r.ocr_end])
    batch_llm_start = _safe_min([r.llm_start for r in records if r.llm_start])
    batch_llm_end   = _safe_max([r.llm_end   for r in records if r.llm_end])

    batch_row = ["BATCH SUMMARY"] + [""] * (len(_AUDIT_FIELDS) * 2 + 1)
    batch_row += [
        _fmt_ts(batch_ocr_start),
        _fmt_ts(batch_ocr_end),
        _fmt_duration(batch_ocr_start, batch_ocr_end),
        _fmt_ts(batch_llm_start),
        _fmt_ts(batch_llm_end),
        _fmt_duration(batch_llm_start, batch_llm_end),
    ]
    ws.append(batch_row)
    row_num = ws.max_row
    for col_idx in range(1, len(batch_row) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.fill   = batch_fill
        cell.font   = batch_font
        cell.border = thin_border

    # ── Auto-fit column widths ────────────────────────────────────────────────
    for col in ws.columns:
        max_len     = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        col_letter  = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    wb.save(str(path))
    logger.info(
        "Audit log saved: %s (%d record(s), accuracy %.2f%%)",
        path, len(records), pct,
    )
    return path


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _compute_matches(extract_result: dict) -> list[bool]:
    """Return a bool list (one entry per _AUDIT_FIELDS) based on comparison data."""
    cmp = extract_result.get("comparison") or {}
    matches = []
    for key, _ in _AUDIT_FIELDS:
        field_cmp = cmp.get(key) or {}
        extracted = str(field_cmp.get("extracted") or "").strip()
        expected  = str(field_cmp.get("expected")  or "").strip()
        matches.append(bool(extracted and expected and extracted == expected))
    return matches


def _build_data_row(record: AuditRecord) -> list:
    """Build the flat list of cell values for one audit row."""
    if record.extract_error or not record.extract_result:
        row = [record.filename] + ["ERROR", ""] * len(_AUDIT_FIELDS) + [0]
    else:
        data = (record.extract_result or {}).get("data") or {}
        cmp  = (record.extract_result or {}).get("comparison") or {}
        row  = [record.filename]
        for key, _ in _AUDIT_FIELDS:
            extracted = str(data.get(key) or "").strip()
            expected  = str((cmp.get(key) or {}).get("expected") or "").strip()
            row += [extracted, expected]
        row.append(sum(_compute_matches(record.extract_result or {})))

    row += [
        _fmt_ts(record.ocr_start),
        _fmt_ts(record.ocr_end),
        _fmt_duration(record.ocr_start, record.ocr_end),
        _fmt_ts(record.llm_start),
        _fmt_ts(record.llm_end),
        _fmt_duration(record.llm_start, record.llm_end),
    ]
    return row


def _safe_min(dts: list[datetime]) -> datetime | None:
    return min(dts) if dts else None


def _safe_max(dts: list[datetime]) -> datetime | None:
    return max(dts) if dts else None